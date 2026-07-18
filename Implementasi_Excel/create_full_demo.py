import openpyxl

def create_full_demo():
    print("Membuka Netflix_Full_Process_V2.xlsx untuk mengambil 20 data pertama...")
    wb_old = openpyxl.load_workbook("Netflix_Full_Process_V2.xlsx", data_only=True)
    ws_old = wb_old["1. Data Mentah"]
    
    reviews = []
    for r in range(2, 22):
        reviews.append(ws_old.cell(row=r, column=3).value)

    print("Membuat file Excel Demo VBA Step-by-Step baru...")
    wb_new = openpyxl.Workbook()
    ws = wb_new.active
    ws.title = "Simulasi Step-By-Step"
    # Header Excel
    headers = [
        "Raw Text", 
        "Case Folding", 
        "Normalization",
        "Remove Noise", 
        "Tokenization", 
        "Stopword Removal", 
        "Stemming",
        "TF-IDF 'terrible'", 
        "TF-IDF 'bad'", 
        "TF-IDF 'worst'",
        "TF-IDF 'hate'",
        "TF-IDF 'good'", 
        "TF-IDF 'love'", 
        "TF-IDF 'great'",
        "TF-IDF 'excellent'",
        "SVM Prediction", 
        "RF Prediction",
        "CLV Lifespan (Bulan)",
        "CLV Potential Loss (Rp)"
    ]
    
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
        
    # Tulis Data dan Rumus Step-by-Step
    for idx, text in enumerate(reviews, start=2):
        r = idx
        
        # A: Raw Text
        ws.cell(row=r, column=1).value = text
        
        # B: Case Folding
        ws.cell(row=r, column=2).value = f"=VBA_CaseFolding(A{r})"
        
        # C: Normalization
        ws.cell(row=r, column=3).value = f"=VBA_Normalization(B{r})"
        
        # D: Remove Noise
        ws.cell(row=r, column=4).value = f"=VBA_RemoveNoise(C{r})"
        
        # E: Tokenization
        ws.cell(row=r, column=5).value = f"=VBA_Tokenize(D{r})"
        
        # F: Stopword
        ws.cell(row=r, column=6).value = f"=VBA_RemoveStopwords(E{r})"
        
        # G: Stemming
        ws.cell(row=r, column=7).value = f"=VBA_Stemming(F{r})"
        
        # H, I, J, K, L, M, N, O: TF-IDF
        ws.cell(row=r, column=8).value = f"=VBA_TFIDF_terribl(G{r})"
        ws.cell(row=r, column=9).value = f"=VBA_TFIDF_bad(G{r})"
        ws.cell(row=r, column=10).value = f"=VBA_TFIDF_worst(G{r})"
        ws.cell(row=r, column=11).value = f"=VBA_TFIDF_hate(G{r})"
        ws.cell(row=r, column=12).value = f"=VBA_TFIDF_good(G{r})"
        ws.cell(row=r, column=13).value = f"=VBA_TFIDF_love(G{r})"
        ws.cell(row=r, column=14).value = f"=VBA_TFIDF_great(G{r})"
        ws.cell(row=r, column=15).value = f"=VBA_TFIDF_excellent(G{r})"
        
        # P: SVM, Q: RF
        ws.cell(row=r, column=16).value = f"=VBA_Predict_SVM(H{r}:O{r})"
        ws.cell(row=r, column=17).value = f"=VBA_Predict_RF(H{r}:O{r})"
        
        # R: CLV Lifespan, S: CLV Loss (menggunakan prediksi SVM karena SVM lebih baik)
        ws.cell(row=r, column=18).value = f"=VBA_CLV_Lifespan(P{r})"
        ws.cell(row=r, column=19).value = f"=VBA_CLV_Loss(R{r})"
        
    wb_new.save("Netflix_StepByStep_Demo.xlsx")
    print("Berhasil membuat Netflix_StepByStep_Demo.xlsx")

if __name__ == "__main__":
    create_full_demo()
