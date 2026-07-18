import openpyxl
from openpyxl.utils import get_column_letter

def add_vba_sheet():
    print("Membuka Netflix_Full_Process_V2.xlsx...")
    wb = openpyxl.load_workbook("Netflix_Full_Process_V2.xlsx")
    
    sheet_name = "5. Simulasi VBA"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
        
    ws = wb.create_sheet(title=sheet_name)
    
    # Headers
    headers = ["Username", "Raw Text", "NLTK Preprocessed (VBA)", "Feature 1", "Feature 2", "Feature 3", "Feature 4", "Random Forest Prediction (VBA)"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
        
    # Copy sample data from Sheet 1
    ws1 = wb["1. Data Mentah"]
    
    print("Menulis formula simulasi VBA...")
    # Add 10 rows for demonstration
    for r in range(2, 12):
        # Username
        ws.cell(row=r, column=1).value = f"='1. Data Mentah'!A{r}"
        # Raw Text
        ws.cell(row=r, column=2).value = f"='1. Data Mentah'!C{r}"
        
        # NLTK Preprocess Formula (VBA)
        ws.cell(row=r, column=3).value = f"=NLTK_Preprocess(B{r})"
        
        # Dummy TF-IDF features just for RF testing
        ws.cell(row=r, column=4).value = 0.6
        ws.cell(row=r, column=5).value = 0.2
        ws.cell(row=r, column=6).value = 0.8
        ws.cell(row=r, column=7).value = 0.1
        
        # Random Forest Formula (VBA)
        ws.cell(row=r, column=8).value = f"=Predict_RandomForest(D{r}:G{r})"

    wb.save("Netflix_Full_Process_V2.xlsx")
    print("Selesai menambahkan Sheet 5!")

if __name__ == "__main__":
    add_vba_sheet()
