import openpyxl

def create_demo():
    print("Membuka Netflix_Full_Process_V2.xlsx untuk mengambil 20 data pertama...")
    wb_old = openpyxl.load_workbook("Netflix_Full_Process_V2.xlsx", data_only=True)
    ws_old = wb_old["1. Data Mentah"]
    
    print("Membuat file Excel Demo VBA baru...")
    wb_new = openpyxl.Workbook()
    ws = wb_new.active
    ws.title = "5. Simulasi VBA"
    
    headers = ["Username", "Raw Text", "NLTK Tokenize & Preprocess", "TF-IDF 'terribl'", "TF-IDF 'bad'", "TF-IDF 'good'", "TF-IDF 'love'", "Random Forest Prediction"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
        
    for r in range(2, 22):
        ws.cell(row=r, column=1).value = ws_old.cell(row=r, column=1).value
        ws.cell(row=r, column=2).value = ws_old.cell(row=r, column=3).value
        
        ws.cell(row=r, column=3).value = f"=NLTK_Preprocess(B{r})"
        
        # TF-IDF Features (VBA Simulation)
        ws.cell(row=r, column=4).value = f'=Calculate_TFIDF(C{r}, "terribl")'
        ws.cell(row=r, column=5).value = f'=Calculate_TFIDF(C{r}, "bad")'
        ws.cell(row=r, column=6).value = f'=Calculate_TFIDF(C{r}, "good")'
        ws.cell(row=r, column=7).value = f'=Calculate_TFIDF(C{r}, "love")'
        
        ws.cell(row=r, column=8).value = f"=Predict_RandomForest(D{r}:G{r})"
        
    wb_new.save("Netflix_Demo_VBA.xlsx")
    print("Berhasil membuat Netflix_Demo_VBA.xlsx")

if __name__ == "__main__":
    create_demo()
