import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()

# --- Sheet 1: Preprocessing ---
ws1 = wb.active
ws1.title = "1. Preprocessing"
ws1.append(['ID', 'Raw Text', 'Cleaned & Stemmed Text'])
data_prep = [
    ['D1', 'being forced to send an email so I can watch a show from a device that is logged into the account before on my home Wi-Fi is ridiculous. You almost killed piracy with this app...', 'forc send email watch show devic log account home wi fi ridicul almost kill piraci applic much better kill applic'],
    ['D2', "won't let me delete my card information, and at random times, it generates a charge to my bank account.", "n t let delet card inform random time gener charg bank account"],
    ['D3', 'Have to completely reset the app and delete all downloads if you accidentally touch the brightness slider. It disables androids built in brightness control.', 'complet reset applic delet download accident touch bright slider disabl android built bright control'],
    ['D4', 'netflix is the best all-in-one entertsinment app. the interface is smooth, and the addition of high quality mobile games...', 'netflix best all in on entertsin applic interfac smooth addit high qualiti mobil game plu tab make easi find download save titl'],
    ['D5', "why is it too often for netflix to suddenly lagging, and I had to sign in so many times it's often failed... I bought it officially from netflix btw", 'often netflix suddenli lag sign mani time often fail need wait never succeed bought offici netflix way'],
    ['D6', "can't run on a Bootloader unlocked device.", "n t run bootload unlock devic"],
    ['D7', 'BRING BACK THE BTS PROFILE PICTURE PLEASE PLEASE PLEASE WE NEED IT ARMY NEEDS IT', 'bring back bt profil pictur pleas pleas pleas need armi need'],
    ['D8', "Rapidly becoming the worst streaming app available. New available content sucks... making it nearly impossible to find anything you actually WANT to watch...", "rapidli becom worst stream applic avail new avail content suck re constantli tri forc bad origin content well alway increas cost justifi differ price tier add advertis constantli chang layout make nearli imposs find anyth actual want watch push thing want watch would give zero star option"],
    ['D9', 'I have no complaints about the quality it’s really good and perfect for a movie marathon. My only concern is that One Piece episodes are always delayed... ', 'no complaint qualiti realli good perfect movi marathon concern one piec episod alway delay arc also miss hope add complet arc futur keep updat prici din siya honest one piec thing wait'],
    ['D10', "I just Abit upset when there is not all movie is got to choose tamil option as audio it's really upsetting hope that Netflix could improve it tqq", 'abit upset not movi got choos tamil option audio realli upset hope netflix could improv tqq']
]
for row in data_prep:
    ws1.append(row)


# --- Sheet 2: Labeling ---
ws2 = wb.create_sheet("2. Labeling")
ws2.append(['ID', 'Score', 'Label Asli (Formula)'])
scores = [1, 1, 1, 5, 2, 1, 5, 1, 2, 4]
for i in range(10):
    row_num = i + 2
    formula = f'=IF(B{row_num}>=4, "Positif", IF(B{row_num}<=2, "Negatif", "Netral"))'
    ws2.append([f'D{i+1}', scores[i], formula])


# --- Sheet 3: TF-IDF ---
ws3 = wb.create_sheet("3. TF-IDF")
ws3.append(['Term', 'DF (Formula)', 'Nilai IDF (Formula)', 'TF di D4 (Formula)', 'Bobot Awal TF*IDF', 'TF-IDF Final D4'])

terms = ['applic', 'netflix', 'watch', 'good', 'movi']
for i, t in enumerate(terms):
    row_num = i + 2
    # DF = COUNTIF('1. Preprocessing'!C$2:C$11, "*"&A2&"*")
    df_form = f'=COUNTIF(\'1. Preprocessing\'!C$2:C$11, "*"&A{row_num}&"*")'
    # IDF = LN((1+$H$2)/(1+B2))+1
    idf_form = f'=LN((1+$H$2)/(1+B{row_num}))+1'
    # TF in D4 = (LEN(D4) - LEN(SUBSTITUTE(D4, Term, ""))) / LEN(Term)
    # D4 cleaned text is in '1. Preprocessing'!C5
    tf_form = f'=(LEN(\'1. Preprocessing\'!$C$5)-LEN(SUBSTITUTE(\'1. Preprocessing\'!$C$5, A{row_num}, "")))/LEN(A{row_num})'
    
    # TF*IDF
    tfidf_init = f'=D{row_num}*C{row_num}'
    # Final TF-IDF normalized
    tfidf_final = f'=E{row_num}/$E$8'
    
    ws3.append([t, df_form, idf_form, tf_form, tfidf_init, tfidf_final])

# Set Total Docs N in H1, H2 AFTER appending to avoid max_row shift
ws3['H1'] = 'Total Doc (N)'
ws3['H2'] = "=COUNTA('1. Preprocessing'!C2:C11)"
ws3['H1'].font = Font(bold=True)

ws3['D8'] = 'L2 Norm (Formula):'
ws3['E8'] = '=SQRT(SUMSQ(E2:E6))'
ws3['D8'].font = Font(bold=True)


# --- Sheet 4: SVM ---
ws4 = wb.create_sheet("4. Klasifikasi SVM")
ws4.append(['Term', 'Bobot Model W', 'Nilai Fitur D4 X', 'Hasil W*X (Formula)'])
svm_weights = [1.0, 1.5, 0.5, 2.0, 1.5]
for i, w in enumerate(svm_weights):
    row_num = i + 2
    ws4.append([terms[i], w, f"='3. TF-IDF'!F{row_num}", f'=B{row_num}*C{row_num}'])

ws4.append(['Bias (b)', 0.5, '', ''])
ws4.append(['TOTAL f(x)', '', '', '=SUM(D2:D6)+B7'])
ws4.append(['Prediksi D4', '', '', '=IF(D8>0, "Positif", "Negatif")'])
ws4['A8'].font = Font(bold=True)


# --- Sheet 5: Evaluasi ---
ws5 = wb.create_sheet("5. Evaluasi")
# We will create a small table for the 10 docs to calculate Confusion Matrix dynamically
ws5.append(['ID', 'Label Asli', 'Prediksi Model (Simulasi)'])
# Let's make mock predictions: 3 TP, 6 TN, 1 FP, 0 FN
mock_preds = ['Negatif', 'Negatif', 'Negatif', 'Positif', 'Negatif', 'Positif', 'Positif', 'Negatif', 'Negatif', 'Positif']

for i in range(10):
    row_num = i + 2
    ws5.append([f'D{i+1}', f"='2. Labeling'!C{row_num}", mock_preds[i]])

# Confusion Matrix Table
ws5['E1'] = 'Confusion Matrix'
ws5['E1'].font = Font(bold=True)

ws5['E2'] = 'True Positive (TP)'
ws5['F2'] = '=COUNTIFS(B2:B11, "Positif", C2:C11, "Positif")'

ws5['E3'] = 'True Negative (TN)'
ws5['F3'] = '=COUNTIFS(B2:B11, "Negatif", C2:C11, "Negatif")'

ws5['E4'] = 'False Positive (FP)'
ws5['F4'] = '=COUNTIFS(B2:B11, "Negatif", C2:C11, "Positif")'

ws5['E5'] = 'False Negative (FN)'
ws5['F5'] = '=COUNTIFS(B2:B11, "Positif", C2:C11, "Negatif")'

# Metrics Table
ws5['H1'] = 'Metrik Evaluasi'
ws5['H1'].font = Font(bold=True)

ws5['H2'] = 'Accuracy'
ws5['I2'] = '=(F2+F3)/SUM(F2:F5)'

ws5['H3'] = 'Precision'
ws5['I3'] = '=F2/(F2+F4)'

ws5['H4'] = 'Recall'
ws5['I4'] = '=F2/(F2+F5)'

ws5['H5'] = 'F1-Score'
ws5['I5'] = '=2*(I3*I4)/(I3+I4)'


wb.save('Simulasi_Manual_Netflix_Full_Formula.xlsx')
print("Full Formula Excel generated successfully.")
