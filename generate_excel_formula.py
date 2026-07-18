import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()

# --- Sheet 1: Preprocessing ---
ws1 = wb.active
ws1.title = "1. Preprocessing"
ws1.append(['ID', 'Raw Text', 'Cleaned & Stemmed Text'])
data_prep = [
    ['D1', 'being forced to send an email so I can watch a show from a device that is logged into the account before on my home Wi-Fi is ridiculous. You almost killed piracy with this app...', 'forc send email watch show devic log account home wi-fi ridicul almost kill piraci applic much better kill applic'],
    ['D2', "won't let me delete my card information, and at random times, it generates a charge to my bank account.", "n't let delet card inform random time gener charg bank account"],
    ['D3', 'Have to completely reset the app and delete all downloads if you accidentally touch the brightness slider. It disables androids built in brightness control.', 'complet reset applic delet download accident touch bright slider disabl android built bright control'],
    ['D4', 'netflix is the best all-in-one entertsinment app. the interface is smooth, and the addition of high quality mobile games...', 'netflix best all-in-on entertsin applic interfac smooth addit high qualiti mobil game plu tab make easi find download save titl'],
    ['D5', "why is it too often for netflix to suddenly lagging, and I had to sign in so many times it's often failed... I bought it officially from netflix btw", 'often netflix suddenli lag sign mani time often fail need wait never succeed bought offici netflix way'],
    ['D6', "can't run on a Bootloader unlocked device.", "n't run bootload unlock devic"],
    ['D7', 'BRING BACK THE BTS PROFILE PICTURE PLEASE PLEASE PLEASE WE NEED IT ARMY NEEDS IT', 'bring back bt profil pictur pleas pleas pleas need armi need'],
    ['D8', "Rapidly becoming the worst streaming app available. New available content sucks... making it nearly impossible to find anything you actually WANT to watch...", "rapidli becom worst stream applic avail new avail content suck 're constantli tri forc bad origin content well alway increas cost justifi differ price tier add advertis constantli chang layout make nearli imposs find anyth actual want watch push thing want watch would give zero star option"],
    ['D9', 'I have no complaints about the quality it’s really good and perfect for a movie marathon. My only concern is that One Piece episodes are always delayed... ', 'no complaint qualiti realli good perfect movi marathon concern one piec episod alway delay arc also miss hope add complet arc futur keep updat prici din siya honest one piec thing wait'],
    ['D10', "I just Abit upset when there is not all movie is got to choose tamil option as audio it's really upsetting hope that Netflix could improve it tqq", 'abit upset not movi got choos tamil option audio realli upset hope netflix could improv tqq']
]
for row in data_prep:
    ws1.append(row)


# --- Sheet 2: Labeling ---
ws2 = wb.create_sheet("2. Labeling")
ws2.append(['ID', 'Score', 'Label Asli (Formula)'])
scores = [1, 1, 1, 5, 2, 1, 5, 1, 2, 4]
for i in range(10):
    row_num = i + 2
    # Formula IF(Score>=4, "Positif", IF(Score<=2, "Negatif", "Netral"))
    formula = f'=IF(B{row_num}>=4, "Positif", IF(B{row_num}<=2, "Negatif", "Netral"))'
    ws2.append([f'D{i+1}', scores[i], formula])


# --- Sheet 3: TF-IDF ---
ws3 = wb.create_sheet("3. TF-IDF")
ws3.append(['Term', 'DF', 'Nilai IDF (Formula)', 'TF (D4)', 'Bobot Awal TF*IDF (Formula)', 'TF-IDF Final D4 (Formula)'])
terms = [
    ['applic', 4, 1],
    ['netflix', 3, 1],
    ['watch', 2, 0],
    ['good', 1, 0],
    ['movi', 2, 0]
]
for i, t in enumerate(terms):
    row_num = i + 2
    # IDF = LN(11 / (1 + DF)) + 1
    idf_formula = f'=LN(11/(1+B{row_num}))+1'
    # TF*IDF
    tfidf_init_formula = f'=D{row_num}*C{row_num}'
    # We will put L2 Norm in cell E8
    final_formula = f'=E{row_num}/$E$8'
    ws3.append([t[0], t[1], idf_formula, t[2], tfidf_init_formula, final_formula])

ws3['D8'] = 'L2 Norm:'
ws3['E8'] = '=SQRT(SUMSQ(E2:E6))'
ws3['D8'].font = Font(bold=True)


# --- Sheet 4: SVM ---
ws4 = wb.create_sheet("4. Klasifikasi SVM")
ws4.append(['Term / Komponen', 'Bobot Model W', 'Nilai Fitur D4 X', 'Hasil W*X (Formula)'])
svm_data = [
    ['applic', 1.0, "='3. TF-IDF'!F2"],  # Link to previous sheet
    ['netflix', 1.5, "='3. TF-IDF'!F3"],
    ['watch', 0.5, "='3. TF-IDF'!F4"],
    ['good', 2.0, "='3. TF-IDF'!F5"],
    ['movi', 1.5, "='3. TF-IDF'!F6"],
]
for i, row in enumerate(svm_data):
    row_num = i + 2
    ws4.append([row[0], row[1], row[2], f'=B{row_num}*C{row_num}'])

# Bias
ws4.append(['Bias (b)', 0.5, '', ''])
# Total
ws4.append(['TOTAL f(x)', '', '', '=SUM(D2:D6)+B7'])
ws4.append(['Prediksi', '', '', '=IF(D8>0, "Positif", "Negatif")'])


# --- Sheet 5: Evaluasi ---
ws5 = wb.create_sheet("5. Evaluasi")
ws5.append(['Confusion Matrix', 'Nilai', '', 'Metrik Evaluasi', 'Hasil (Formula)'])
ws5.append(['True Positive (TP)', 4, '', 'Accuracy', '=(B2+B3)/SUM(B2:B5)'])
ws5.append(['True Negative (TN)', 4, '', 'Precision', '=B2/(B2+B4)'])
ws5.append(['False Positive (FP)', 1, '', 'Recall', '=B2/(B2+B5)'])
ws5.append(['False Negative (FN)', 1, '', 'F1-Score', '=2*(E3*E4)/(E3+E4)'])


wb.save('Simulasi_Manual_Netflix_Formula.xlsx')
print("Formula Excel generated successfully.")
